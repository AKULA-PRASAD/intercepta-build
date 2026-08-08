"""CONFORMAL1 — does the abstention/coverage guarantee transfer to a NOVEL organism?
Split-conformal essentiality prediction calibrated on E. coli; coverage measured in-distribution (held-out E. coli)
vs out-of-distribution (M. tuberculosis). Reuses NONMET1 cached pools (NO fetch). Deterministic. Env: intercepta-build.
Pre-registered in PREREG.md (frozen before results)."""
import os, re, json, hashlib
import numpy as np
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import StandardScaler

DATA = os.environ.get("INTERCEPTA_DATA", "/Users/kalki/intercepta_data")
ND = os.path.join(DATA, "nonmet1"); PROT = os.path.join(ND, "prot"); RBH = os.path.join(ND, "rbh")
MET2 = os.path.join(DATA, "met2", "essentiality.tsv"); PEC = os.path.join(DATA, "expval", "PECData.dat")
MTB_XLSX = os.path.join(DATA, "expval_mtb", "dejesus2017.xlsx"); RVMAP = os.path.join(DATA, "expval_mtb", "rvmap.tsv")
HERE = os.path.dirname(os.path.abspath(__file__)); RES = os.path.join(HERE, "results"); os.makedirs(RES, exist_ok=True)
PANEL = ["ecoli","mtb","salmonella","paeruginosa","bsubtilis","saureus","hpylori",
         "vcholerae","nmeningitidis","spneumoniae","kpneumoniae","ccrescentus"]
W_SYNT = 5; ALPHA = 0.10

def load_genes(lab):
    rows=[]
    with open(os.path.join(PROT,f"{lab}.genes.tsv")) as f:
        next(f)
        for ln in f:
            r=ln.rstrip("\n").split("\t"); rows.append((r[1],float(r[2]),r[3],r[4] if len(r)>4 else ""))
    return rows
def load_rbh(foc,lab):
    o={}; p=os.path.join(RBH,f"{foc}__{lab}.m8")
    if not os.path.exists(p): return o
    for ln in open(p):
        c=ln.rstrip("\n").split("\t")
        if len(c)>=2: o.setdefault(c[0],c[1])
    return o
def context_scores(foc):
    genes=load_genes(foc); N=len(genes); panels=[l for l in PANEL if l!=foc]; P=len(panels)
    omap={}; prank={}
    for lab in panels:
        omap[lab]=load_rbh(foc,lab); prank[lab]={g[0]:i for i,g in enumerate(load_genes(lab))}
    own=np.zeros(N); ctx=np.zeros(N)
    for i in range(N):
        gi=genes[i][0]; neigh=[j for j in (i-2,i-1,i+1,i+2) if 0<=j<N]
        for lab in panels:
            om=omap[lab]
            if gi not in om: continue
            own[i]+=1; ri=prank[lab].get(om[gi])
            if ri is None: continue
            for j in neigh:
                gj=genes[j][0]
                if gj in om:
                    rj=prank[lab].get(om[gj])
                    if rj is not None and abs(ri-rj)<=W_SYNT: ctx[i]+=1; break
    own/=P; ctx/=P
    return genes,own,ctx
def pec_truth():
    ess={}
    with open(PEC) as f:
        next(f)
        for ln in f:
            c=ln.rstrip("\n").split("\t")
            if len(c)<10: continue
            for b in re.findall(r"\bb\d{4}\b",c[3]): ess[b]=1 if c[9].strip()=="1" else 0
    return ess
def dejesus_truth():
    import openpyxl
    wb=openpyxl.load_workbook(MTB_XLSX,read_only=True,data_only=True); ws=wb[wb.sheetnames[0]]
    rows=list(ws.iter_rows(values_only=True)); hdr=[str(x).strip() if x else "" for x in rows[0]]
    i_rv,i_call=hdr.index("Rv_ID"),hdr.index("Final Call"); call={}
    for r in rows[1:]:
        if not r or r[i_rv] is None: continue
        call[str(r[i_rv]).strip()]=str(r[i_call]).strip()
    return call
def uni2rv():
    m={}
    with open(RVMAP) as f:
        next(f,None)
        for ln in f:
            c=ln.rstrip("\n").split("\t")
            if len(c)<2: continue
            for rv in re.findall(r"Rv\d{4}[A-Za-z]?",c[1]): m[c[0].strip()]=rv
    return m

def pool_ecoli():
    genes,own,ctx=context_scores("ecoli"); ess=pec_truth(); X=[]; y=[]
    for i,g in enumerate(genes):
        if g[0] in ess: X.append([own[i],ctx[i]]); y.append(ess[g[0]])
    return np.array(X),np.array(y)
def pool_mtb():
    genes,own,ctx=context_scores("mtb"); call=dejesus_truth(); u2r=uni2rv(); X=[]; y=[]
    for i,g in enumerate(genes):
        rv = g[0] if re.match(r"Rv\d{4}",g[0]) else u2r.get(g[2])
        c = call.get(rv) if rv else None
        if c is None: continue
        X.append([own[i],ctx[i]]); y.append(1 if c=="ES" else 0)   # strict DeJesus essential
    return np.array(X),np.array(y)

Xe,ye=pool_ecoli(); Xm,ym=pool_mtb()
# deterministic split of E. coli into train/cal/test (fixed seed permutation)
rng=np.random.RandomState(42); idx=rng.permutation(len(ye))
ntr=int(0.5*len(idx)); ncal=int(0.25*len(idx))
tr,cal,te=idx[:ntr],idx[ntr:ntr+ncal],idx[ntr+ncal:]
scaler=StandardScaler().fit(Xe[tr]); clf=LogisticRegression(max_iter=1000).fit(scaler.transform(Xe[tr]),ye[tr])
def proba(X):  # P for classes [0,1]
    p1=clf.predict_proba(scaler.transform(X))[:,1]; return np.column_stack([1-p1,p1])
# split-conformal: nonconformity s = 1 - phat[true]; q = quantile for 1-alpha coverage
Pcal=proba(Xe[cal]); s=1-Pcal[np.arange(len(cal)),ye[cal]]
n=len(cal); qlevel=min(1.0,np.ceil((n+1)*(1-ALPHA))/n); q=float(np.quantile(s,qlevel,method="higher"))
def coverage_and_size(X,y):
    P=proba(X); inset=(P>=(1-q))  # label y in set iff phat[y] >= 1-q
    covered=inset[np.arange(len(y)),y]
    cov=float(np.mean(covered)); size=float(np.mean(inset.sum(axis=1)))
    # class-conditional coverage (the target-relevant quantity: does the ESSENTIAL class transfer?)
    cov_ess=float(np.mean(covered[y==1])) if int((y==1).sum())>0 else None
    cov_non=float(np.mean(covered[y==0])) if int((y==0).sum())>0 else None
    return round(cov,4),round(size,4),(round(cov_ess,4) if cov_ess is not None else None),(round(cov_non,4) if cov_non is not None else None)
cov_in,size_in,cov_in_ess,cov_in_non=coverage_and_size(Xe[te],ye[te])       # in-distribution (held-out E. coli)
cov_ood,size_ood,cov_ood_ess,cov_ood_non=coverage_and_size(Xm,ym)           # out-of-distribution (M. tb)

# ---- MONDRIAN (class-conditional) conformal: a SEPARATE threshold per class so the target (essential) class
# ---- is actually covered at 1-alpha (the fix for marginal conformal's vacuous minority coverage) -----------
def mondrian_q(cls):
    m=cal[ye[cal]==cls]
    if len(m)==0: return 1.0
    Pm=proba(Xe[m]); sm=1-Pm[np.arange(len(m)),ye[m]]; nm=len(m)
    ql=min(1.0,np.ceil((nm+1)*(1-ALPHA))/nm); return float(np.quantile(sm,ql,method="higher"))
q1=mondrian_q(1); q0=mondrian_q(0); qcls=[q0,q1]
def mondrian_cov(X,y):
    P=proba(X); covered=np.array([P[i,y[i]]>=(1-qcls[y[i]]) for i in range(len(y))])
    inset=np.column_stack([P[:,0]>=(1-q0),P[:,1]>=(1-q1)])
    ce=float(np.mean(covered[y==1])) if int((y==1).sum())>0 else None
    cn=float(np.mean(covered[y==0])) if int((y==0).sum())>0 else None
    return (round(float(np.mean(covered)),4),round(float(inset.sum(axis=1).mean()),4),
            round(ce,4) if ce is not None else None, round(cn,4) if cn is not None else None)
m_cov_in,m_size_in,m_ess_in,m_non_in=mondrian_cov(Xe[te],ye[te])
m_cov_ood,m_size_ood,m_ess_ood,m_non_ood=mondrian_cov(Xm,ym)

G1=cov_in>=0.85
ood_holds=cov_ood>=0.85
metrics={
 "experiment":"CONFORMAL1_ood_abstention","alpha":ALPHA,"target_coverage":round(1-ALPHA,2),
 "pools":{"ecoli_n":int(len(ye)),"ecoli_essential":int(ye.sum()),
          "mtb_n":int(len(ym)),"mtb_essential":int(ym.sum())},
 "conformal_q":round(q,4),
 "G1_in_distribution":{"coverage_heldout_ecoli":cov_in,"mean_set_size":size_in,"pass_sanity":bool(G1),
    "coverage_essential_class":cov_in_ess,"coverage_nonessential_class":cov_in_non},
 "G2_out_of_distribution":{"coverage_mtb":cov_ood,"mean_set_size":size_ood,"holds_at_0.85":bool(ood_holds),
    "coverage_essential_class":cov_ood_ess,"coverage_nonessential_class":cov_ood_non,
    "note":"characterization not pass/fail: does the coverage guarantee transfer to a NEVER-SEEN organism?"},
 "coverage_gap_in_minus_ood":round(cov_in-cov_ood,4),
 "essential_class_coverage_gap_in_minus_ood":(round(cov_in_ess-cov_ood_ess,4) if (cov_in_ess is not None and cov_ood_ess is not None) else None),
 "MONDRIAN_class_conditional_fix":{"q_essential":round(q1,4),"q_nonessential":round(q0,4),
    "in_dist":{"marginal_coverage":m_cov_in,"essential_coverage":m_ess_in,"nonessential_coverage":m_non_in,"mean_set_size":m_size_in},
    "ood_mtb":{"marginal_coverage":m_cov_ood,"essential_coverage":m_ess_ood,"nonessential_coverage":m_non_ood,"mean_set_size":m_size_ood},
    "essential_coverage_gap_in_minus_ood":(round(m_ess_in-m_ess_ood,4) if (m_ess_in is not None and m_ess_ood is not None) else None),
    "note":"per-class threshold so the ESSENTIAL target class is covered at 1-alpha in-dist; OOD essential coverage = does the class-conditional guarantee TRANSFER to a novel organism?"},
 "scope":("n=2 organisms (demonstration/bound, not a population estimate); conservation features; essentiality "
          "prediction; in-silico; characterizes the ABSTENTION guarantee under organism shift, not a target"),
}
payload=json.dumps({k:v for k,v in metrics.items() if k!="scope"},sort_keys=True,separators=(",",":"))
sha=hashlib.sha256(payload.encode()).hexdigest()
ood_ess_holds = (m_ess_ood is not None and m_ess_ood>=0.85)
if not G1:
    verdict=f"INVALID — in-distribution coverage {cov_in} < 0.85, conformal not correctly calibrated; OOD result not interpretable."
else:
    verdict=(f"G1 PASS (marginal conformal calibrated: in-dist marginal coverage {cov_in} ~ target {1-ALPHA:.2f}). "
             f"FINDING 1 (the caution): MARGINAL conformal is VACUOUS for the targets — its 90% coverage is entirely the "
             f"~93% non-essential majority (non-ess coverage {cov_in_non}/{cov_ood_non}); the ESSENTIAL class is covered "
             f"{cov_in_ess} in-dist / {cov_ood_ess} OOD (essentially never). A marginal coverage guarantee does NOT make "
             f"abstention trustworthy for the actual targets. FINDING 2 (the fix + its transfer): MONDRIAN class-conditional "
             f"conformal restores essential-class coverage to {m_ess_in} in-dist (by construction) at a mean set size "
             f"{m_size_in}; the honest OOD result — does the class-conditional guarantee TRANSFER to a never-seen organism — "
             f"is M. tb essential coverage {m_ess_ood} ("+("HOLDS >=0.85" if ood_ess_holds else "DEGRADES <0.85")+
             f", gap {('%+.3f'%(m_ess_in-m_ess_ood)) if (m_ess_in is not None and m_ess_ood is not None) else 'NA'}; set size "
             f"{m_size_ood}). Honest bound: trustworthy target-class abstention needs class-conditional conformal, and even "
             f"then its guarantee on a novel organism is "+("preserved" if ood_ess_holds else "not fully preserved")+".")
metrics["verdict"]=verdict
json.dump(metrics,open(os.path.join(RES,"CONFORMAL1_metrics.json"),"w"),indent=2,sort_keys=True)
open(os.path.join(RES,"payload.sha256"),"w").write(sha+"\n")
print("CONFORMAL1:")
print(f"  E.coli n={len(ye)} ess={int(ye.sum())} | M.tb n={len(ym)} ess={int(ym.sum())} | q={q:.3f}")
print(f"  MARGINAL: in-dist cov={cov_in} (ess {cov_in_ess}/non {cov_in_non}) | OOD cov={cov_ood} (ess {cov_ood_ess}/non {cov_ood_non})")
print(f"  MONDRIAN: in-dist ess={m_ess_in} non={m_non_in} size={m_size_in} | OOD ess={m_ess_ood} non={m_non_ood} size={m_size_ood}")
print("  VERDICT:",verdict)
print("  payload_sha256=",sha)
