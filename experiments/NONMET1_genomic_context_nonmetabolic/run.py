"""NONMET1 analysis (reproduced x2 byte-identical). Tests the pre-registered hypothesis (PREREG.md): does
CONSERVED GENOMIC CONTEXT (synteny breadth) enrich for EXPERIMENTAL essentiality on the NON-METABOLIC subproteome,
and does it add BEYOND the raw sequence-conservation-breadth null? Consumes prep.py caches in $INTERCEPTA_DATA/nonmet1/.
Deterministic: no RNG, StratifiedKFold(shuffle=False), mmseqs cached. Env: intercepta-build."""
import os, re, json, time, hashlib
import numpy as np
from scipy.stats import fisher_exact, spearmanr, pearsonr
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import roc_auc_score

DATA = os.environ.get("INTERCEPTA_DATA", "/Users/kalki/intercepta_data")
ND = os.path.join(DATA, "nonmet1"); PROT = os.path.join(ND, "prot"); RBH = os.path.join(ND, "rbh")
MET2 = os.path.join(DATA, "met2", "essentiality.tsv")
PEC = os.path.join(DATA, "expval", "PECData.dat")
MTB_XLSX = os.path.join(DATA, "expval_mtb", "dejesus2017.xlsx")
RVMAP = os.path.join(DATA, "expval_mtb", "rvmap.tsv")
HERE = os.path.dirname(os.path.abspath(__file__))
PANEL = ["ecoli","mtb","salmonella","paeruginosa","bsubtilis","saureus","hpylori",
         "vcholerae","nmeningitidis","spneumoniae","kpneumoniae","ccrescentus"]
K_NEIGH = 4      # 2 up + 2 down  (locked)
W_SYNT = 5       # synteny gene-rank window (locked)


def load_genes(lab):
    """rank-ordered list of (locus_tag, midpoint, uniprot, gene); rank == list index."""
    rows = []
    with open(os.path.join(PROT, f"{lab}.genes.tsv")) as f:
        next(f)
        for ln in f:
            r = ln.rstrip("\n").split("\t")
            rows.append((r[1], float(r[2]), r[3], r[4] if len(r) > 4 else ""))
    return rows


def load_rbh(foc, lab):
    """focal_locus -> panel_locus (RBH). easy-rbh gives reciprocal best pairs."""
    o = {}
    p = os.path.join(RBH, f"{foc}__{lab}.m8")
    if not os.path.exists(p):
        return o
    for ln in open(p):
        c = ln.rstrip("\n").split("\t")
        if len(c) < 2:
            continue
        o.setdefault(c[0], c[1])   # first (best) pair per query
    return o


def context_scores(foc):
    genes = load_genes(foc)
    N = len(genes)
    tag2rank = {g[0]: i for i, g in enumerate(genes)}
    panels = [l for l in PANEL if l != foc]
    P = len(panels)
    # ortholog maps + panel ranks
    omap = {}; prank = {}
    for lab in panels:
        omap[lab] = load_rbh(foc, lab)
        prank[lab] = {g[0]: i for i, g in enumerate(load_genes(lab))}
    own = np.zeros(N); ctx = np.zeros(N)
    for i in range(N):
        gi = genes[i][0]
        neigh = [j for j in (i-2, i-1, i+1, i+2) if 0 <= j < N]
        for lab in panels:
            om = omap[lab]
            if gi not in om:
                continue
            own[i] += 1
            oi = om[gi]; ri = prank[lab].get(oi)
            if ri is None:
                continue
            syn = False
            for j in neigh:
                gj = genes[j][0]
                if gj in om:
                    rj = prank[lab].get(om[gj])
                    if rj is not None and abs(ri - rj) <= W_SYNT:
                        syn = True; break
            if syn:
                ctx[i] += 1
    own /= P; ctx /= P
    cond = np.where(own > 0, ctx / np.maximum(own, 1e-9), 0.0)
    return genes, own, ctx, cond


def metabolic_set_ecoli():
    s = set()
    for ln in open(MET2):
        p = ln.rstrip().split("\t")
        if p[0] == "ecoli":
            s.add(p[1])       # uniprot
    return s


def metabolic_set_mtb():
    s = set()
    for ln in open(MET2):
        p = ln.rstrip().split("\t")
        if p[0] == "mtb":
            s.add(p[1])
    return s


def pec_truth():
    """b-number -> (essential 0/1, n_pmid)."""
    ess = {}; pmid = {}
    with open(PEC) as f:
        next(f)
        for ln in f:
            c = ln.rstrip("\n").split("\t")
            if len(c) < 10:
                continue
            bs = re.findall(r"\bb\d{4}\b", c[3])
            e = 1 if c[9].strip() == "1" else 0
            npm = len([x for x in re.split(r"[,\s]+", c[12].strip()) if x]) if len(c) > 12 else 0
            for b in bs:
                ess[b] = e; pmid[b] = npm
    return ess, pmid


def dejesus_truth():
    import openpyxl
    wb = openpyxl.load_workbook(MTB_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(x).strip() if x else "" for x in rows[0]]
    i_rv, i_call = hdr.index("Rv_ID"), hdr.index("Final Call")
    call = {}
    for r in rows[1:]:
        if not r or r[i_rv] is None:
            continue
        call[str(r[i_rv]).strip()] = str(r[i_call]).strip()
    return call


def rv2uniprot():
    m = {}
    with open(RVMAP) as f:
        next(f, None)
        for ln in f:
            c = ln.rstrip("\n").split("\t")
            if len(c) < 2:
                continue
            for rv in re.findall(r"Rv\d{4}[A-Za-z]?", c[1]):
                m.setdefault(rv, c[0].strip())
    return m


def cv_auroc(X, y):
    """pooled out-of-fold AUROC, 5-fold stratified, per-fold standardized, deterministic."""
    X = np.asarray(X, float); y = np.asarray(y, int)
    oof = np.zeros(len(y))
    skf = StratifiedKFold(n_splits=5, shuffle=False)
    for tr, te in skf.split(X, y):
        sc = StandardScaler().fit(X[tr])
        clf = LogisticRegression(C=1.0, max_iter=1000, solver="lbfgs")
        clf.fit(sc.transform(X[tr]), y[tr])
        oof[te] = clf.predict_proba(sc.transform(X[te]))[:, 1]
    return float(roc_auc_score(y, oof))


def enrichment(ctx, y):
    thr = float(np.median(ctx))
    hi = ctx >= thr
    a = int(np.sum(hi & (y == 1))); b = int(np.sum(hi & (y == 0)))
    c = int(np.sum(~hi & (y == 1))); d = int(np.sum(~hi & (y == 0)))
    orr, p = fisher_exact([[a, b], [c, d]], alternative="greater")
    return thr, {"both": a, "hi_noness": b, "lo_ess": c, "neither": d}, float(orr), float(p)


def analyze_ecoli():
    genes, own, ctx, cond = context_scores("ecoli")
    met = metabolic_set_ecoli()
    ess, pmid = pec_truth()
    rows = []
    for i, (tag, mid, up, sym) in enumerate(genes):
        is_met = up in met if up else False
        if is_met:
            continue                       # NON-METABOLIC subproteome only
        if tag not in ess:
            continue                       # require a PEC essentiality call
        rows.append((tag, own[i], ctx[i], cond[i], ess[tag], np.log1p(pmid.get(tag, 0))))
    tags = [r[0] for r in rows]
    OWN = np.array([r[1] for r in rows]); CTX = np.array([r[2] for r in rows])
    COND = np.array([r[3] for r in rows]); Y = np.array([r[4] for r in rows]); STU = np.array([r[5] for r in rows])
    return run_stats("E_coli", "PEC_class1", tags, OWN, CTX, COND, Y, STU)


def analyze_mtb():
    genes, own, ctx, cond = context_scores("mtb")
    met_up = metabolic_set_mtb()
    r2u = rv2uniprot()
    call = dejesus_truth()
    rows = []
    for i, (tag, mid, up, sym) in enumerate(genes):
        uni = r2u.get(tag, "")
        is_met = uni in met_up if uni else False
        if is_met:
            continue
        if tag not in call:
            continue                       # require a DeJesus call
        y = 1 if call[tag] == "ES" else 0
        rows.append((tag, own[i], ctx[i], cond[i], y))
    tags = [r[0] for r in rows]
    OWN = np.array([r[1] for r in rows]); CTX = np.array([r[2] for r in rows])
    COND = np.array([r[3] for r in rows]); Y = np.array([r[4] for r in rows])
    return run_stats("M_tuberculosis", "DeJesus2017_ES", tags, OWN, CTX, COND, Y, None)


def run_stats(org, truth, tags, OWN, CTX, COND, Y, STU):
    n = len(Y); npos = int(Y.sum())
    r_own_ctx = float(pearsonr(OWN, CTX)[0]) if n > 2 else float("nan")
    sr_own_ctx = float(spearmanr(OWN, CTX)[0]) if n > 2 else float("nan")
    au_m1 = cv_auroc(OWN.reshape(-1, 1), Y)
    au_m2 = cv_auroc(np.column_stack([OWN, CTX]), Y)
    au_m2b = cv_auroc(np.column_stack([OWN, COND]), Y)
    d_auroc = au_m2 - au_m1
    d_auroc_cond = au_m2b - au_m1
    # full-fit standardized coefficient of context in M2
    sc = StandardScaler().fit(np.column_stack([OWN, CTX]))
    clf = LogisticRegression(C=1.0, max_iter=1000, solver="lbfgs")
    clf.fit(sc.transform(np.column_stack([OWN, CTX])), Y)
    coef_own, coef_ctx = float(clf.coef_[0][0]), float(clf.coef_[0][1])
    thr, cont, orr, p_fish = enrichment(CTX, Y)
    out = {
        "organism": org, "truth_source": truth,
        "n_nonmetabolic_tested": n, "n_experimental_essential": npos,
        "prevalence": round(npos / n, 6) if n else None,
        "pearson_own_vs_context": round(r_own_ctx, 6), "spearman_own_vs_context": round(sr_own_ctx, 6),
        "auroc_M1_own_only": round(au_m1, 6),
        "auroc_M2_own_plus_context": round(au_m2, 6),
        "delta_auroc_context_beyond_own": round(d_auroc, 6),
        "auroc_M2b_own_plus_conditional_synteny": round(au_m2b, 6),
        "delta_auroc_conditional_beyond_own": round(d_auroc_cond, 6),
        "stdcoef_own_in_M2": round(coef_own, 6), "stdcoef_context_in_M2": round(coef_ctx, 6),
        "enrichment_context_median_thr": round(thr, 6),
        "enrichment_contingency": cont,
        "enrichment_odds_ratio": round(orr, 6),
        "enrichment_fisher_p": round(p_fish, 10),
    }
    if STU is not None:
        au_m3 = cv_auroc(np.column_stack([OWN, STU]), Y)
        au_m4 = cv_auroc(np.column_stack([OWN, STU, CTX]), Y)
        out["studybias_auroc_M3_own_plus_study"] = round(au_m3, 6)
        out["studybias_auroc_M4_own_study_context"] = round(au_m4, 6)
        out["studybias_delta_auroc_context_beyond_own_study"] = round(au_m4 - au_m3, 6)
        out["study_proxy_pearson_with_context"] = round(float(pearsonr(STU, CTX)[0]), 6)
    return out


def gate_ecoli(m):
    passA = m["delta_auroc_context_beyond_own"] >= 0.03
    passB = (m["enrichment_odds_ratio"] > 2.0) and (m["enrichment_fisher_p"] < 0.01)
    sb = m.get("studybias_delta_auroc_context_beyond_own_study", None)
    sb_ok = (sb is None) or (sb >= 0.02)
    passed = bool(passA and passB)
    return passed, passA, passB, (passed and sb_ok)


def main():
    t0 = time.time()
    res = {"ecoli": analyze_ecoli(), "mtb": analyze_mtb()}
    passed, passA, passB, sb_ok = gate_ecoli(res["ecoli"])
    payload = {"experiment": "NONMET1_genomic_context_nonmetabolic",
               "hypothesis": "conserved genomic CONTEXT (synteny breadth) enriches non-metabolic essentiality "
                             "and adds BEYOND raw sequence-conservation breadth",
               "params": {"panel_size": len(PANEL), "k_neighborhood": K_NEIGH, "W_synteny_window": W_SYNT,
                          "rbh_min_seq_id": 0.30, "rbh_cov": 0.5, "cov_mode": 0, "evalue": 1e-5,
                          "cv": "StratifiedKFold_5_shuffleFalse", "logreg_C": 1.0},
               "gate": {"require_delta_auroc_ge": 0.03, "require_enrichment_OR_gt": 2.0, "require_fisher_p_lt": 0.01,
                        "studybias_delta_auroc_ge": 0.02},
               "results": res,
               "gate_eval_ecoli": {"passA_delta_auroc": bool(passA), "passB_enrichment": bool(passB),
                                   "PASS_primary": bool(passed), "PASS_studybias_robust": bool(sb_ok)}}
    # SHA over sorted-key JSON EXCLUDING verdict + provenance
    core = json.dumps(payload, sort_keys=True, separators=(",", ":"))
    sha = hashlib.sha256(core.encode()).hexdigest()
    m = res["ecoli"]
    if passed:
        verdict = (f"PASS — conserved genomic context adds beyond the conservation null on the E. coli non-metabolic "
                   f"subproteome (ΔAUROC={m['delta_auroc_context_beyond_own']:+.3f}, enrichment OR={m['enrichment_odds_ratio']:.2f} "
                   f"p={m['enrichment_fisher_p']:.1e}).")
    else:
        verdict = (f"FAIL (first-class NEGATIVE) — genomic-context/synteny conservation does NOT add a decisive signal "
                   f"beyond raw sequence-conservation breadth on the E. coli non-metabolic subproteome "
                   f"(ΔAUROC={m['delta_auroc_context_beyond_own']:+.3f} vs gate +0.030; enrichment OR={m['enrichment_odds_ratio']:.2f}, "
                   f"p={m['enrichment_fisher_p']:.1e}). Own-vs-context Pearson r={m['pearson_own_vs_context']:.2f} "
                   f"(largely collinear). Closes the 'genomic context cracks the non-metabolic half' door, as MET4 closed PPI.")
    payload["verdict"] = verdict
    payload["provenance"] = {"generated": time.strftime("%Y-%m-%dT%H:%M:%S"), "runtime_s": round(time.time() - t0, 1),
                             "data": ND}
    os.makedirs(os.path.join(HERE, "results"), exist_ok=True)
    with open(os.path.join(HERE, "results", "NONMET1_metrics.json"), "w") as f:
        json.dump(payload, f, sort_keys=True, indent=2)
    with open(os.path.join(HERE, "results", "payload.sha256"), "w") as f:
        f.write(sha + "\n")
    print(json.dumps(res, indent=2, sort_keys=True))
    print("\nPAYLOAD_SHA256:", sha)
    print("VERDICT:", verdict)


if __name__ == "__main__":
    main()
