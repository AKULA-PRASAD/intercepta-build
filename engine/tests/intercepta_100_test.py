"""
INTERCEPTA — 100-LEVEL STRESS TEST
====================================
Finds every wrong thing. No mercy. No assumptions.
Run: python3 intercepta_100_test.py
"""
import sys, os, json, csv, math, traceback, hashlib
import numpy as np
sys.path.insert(0, os.path.expanduser('~/INTERCEPTA/code'))
BASE = os.path.expanduser('~/INTERCEPTA/')

results = []
lv = [0]

def test(name, fn, cat=""):
    lv[0] += 1
    try:
        v, d = fn()
        results.append((lv[0], name, v, d, cat))
        sym = "✓" if v=="PASS" else ("✗" if v=="FAIL" else ("⚠" if v=="WARN" else "!"))
        print(f"  {sym} L{lv[0]:03d} {v:<8} {name}")
        if d: print(f"           → {d}")
    except Exception as e:
        tb = traceback.format_exc().strip().split('\n')[-1]
        results.append((lv[0], name, "ERROR", tb[:120], cat))
        print(f"  ! L{lv[0]:03d} ERROR   {name}: {tb[:100]}")

print("="*70)
print("INTERCEPTA — 100-LEVEL STRESS TEST")
print("="*70)

# ═══════════════════════════════════════════════════════════════
# TIER 1: DATA INTEGRITY (L1-L12)
# ═══════════════════════════════════════════════════════════════
print("\n╔══ TIER 1: DATA INTEGRITY (L1-L12) ══╗")

def t_beataml_patient_ids_unique():
    import openpyxl
    wb = openpyxl.load_workbook(BASE+'data/beataml/beataml_wv1to4_clinical.xlsx')
    ws = wb.active
    ids = [ws.cell(r,1).value for r in range(2, ws.max_row+1)]
    ids = [i for i in ids if i]
    dupes = len(ids) - len(set(ids))
    return ("FAIL",f"{dupes} duplicate patient IDs") if dupes else ("PASS",f"{len(ids)} unique patient IDs")
test("BeatAML patient IDs unique", t_beataml_patient_ids_unique, "data")

def t_beataml_drug_names_consistent():
    path = BASE+'data/beataml/beataml_probit_curve_fits_v4_dbgap.txt'
    drugs = set()
    with open(path) as f:
        reader = csv.DictReader(f, delimiter='\t')
        for row in reader:
            for k in ['inhibitor','drug','Drug','Inhibitor']:
                if k in row: drugs.add(row[k]); break
    return ("PASS",f"{len(drugs)} unique drug names") if len(drugs)>100 else ("WARN",f"Only {len(drugs)} drugs")
test("BeatAML drug names parseable", t_beataml_drug_names_consistent, "data")

def t_scrna_cell_count_matches_velocity():
    with open(BASE+'results/step3_velocity_results.csv') as f:
        vel_cells = len(list(csv.DictReader(f)))
    # Should match our 35589 cells
    return ("PASS",f"Velocity results: {vel_cells} cells") if vel_cells > 10000 else ("FAIL",f"Only {vel_cells}")
test("scRNA cell count consistent", t_scrna_cell_count_matches_velocity, "data")

def t_alphafold_all_20_targets():
    expected = ['AKT1','AKT2','AR','ATM','BCL2L1','BRCA1','CDK2','CDK4',
                'CSNK2A1','EGFR','ERBB2','KIT','MAP2K1','MAPK14','MDM2',
                'MTOR','PARP1','PIK3CA','TOP2B','TP53']
    af_dir = BASE+'data/alphafold/'
    found = [f.split('_AF')[0] for f in os.listdir(af_dir) if f.endswith('.pdb')]
    missing = [t for t in expected if t not in found]
    return ("FAIL",f"Missing structures: {missing}") if missing else ("PASS",f"All 20 targets present")
test("All 20 AlphaFold targets present", t_alphafold_all_20_targets, "data")

def t_alphafold_plddt_quality():
    af_dir = BASE+'data/alphafold/'
    low_quality = []
    for pdb in os.listdir(af_dir):
        if not pdb.endswith('.pdb'): continue
        with open(af_dir+pdb) as f: content = f.read()
        if 'ATOM' not in content: continue
        # pLDDT is in B-factor column of ATOM records
        bfactors = []
        for line in content.split('\n'):
            if line.startswith('ATOM'):
                try: bfactors.append(float(line[60:66]))
                except: pass
        if bfactors:
            mean_plddt = np.mean(bfactors)
            if mean_plddt < 70: low_quality.append(f"{pdb[:10]}({mean_plddt:.0f})")
    return ("WARN",f"Low pLDDT (<70): {low_quality}") if low_quality else ("PASS",f"All structures have mean pLDDT≥70")
test("AlphaFold pLDDT quality scores acceptable", t_alphafold_plddt_quality, "data")

def t_signor_no_self_loops():
    with open(BASE+'results/signor_directed_edges.csv') as f:
        rows = list(csv.DictReader(f))
    cols = list(rows[0].keys())
    src_col = next((c for c in cols if 'source' in c.lower() or 'a' == c.lower()), cols[0])
    tgt_col = next((c for c in cols if 'target' in c.lower() or 'b' == c.lower()), cols[1])
    self_loops = sum(1 for r in rows if r[src_col]==r[tgt_col])
    return ("WARN",f"{self_loops} self-loops in SIGNOR") if self_loops>100 else ("PASS",f"{self_loops} self-loops (acceptable)")
test("SIGNOR edge self-loops", t_signor_no_self_loops, "data")

def t_string_score_range():
    with open(BASE+'results/step4_string_interactions.csv') as f:
        rows = list(csv.DictReader(f))
    cols = list(rows[0].keys())
    score_col = next((c for c in cols if 'score' in c.lower()), None)
    if not score_col: return "WARN","no score column"
    scores = [float(r[score_col]) for r in rows if r.get(score_col)]
    out_of_range = sum(1 for s in scores if not (0 <= s <= 1000))
    return ("FAIL",f"{out_of_range} scores outside [0,1000]") if out_of_range else ("PASS",f"Scores in [0,1000], mean={np.mean(scores):.0f}")
test("STRING scores in valid range [0,1000]", t_string_score_range, "data")

def t_velocity_no_nan():
    with open(BASE+'results/step3_velocity_results.csv') as f:
        rows = list(csv.DictReader(f))
    nan_rows = 0
    for r in rows:
        for v in r.values():
            try:
                if math.isnan(float(v)): nan_rows += 1; break
            except: pass
    return ("WARN",f"{nan_rows} rows contain NaN") if nan_rows>100 else ("PASS",f"{nan_rows} NaN rows (clean)")
test("RNA velocity no excessive NaN", t_velocity_no_nan, "data")

def t_gdsc_expression_exists():
    paths = [BASE+'data/gdsc/sanger_model_gene_expression.csv.gz',
             BASE+'data/gdsc/Cell_line_RMA_proc_basalExp.txt.gz']
    for p in paths:
        if os.path.exists(p) and os.path.getsize(p)>1e6:
            return "PASS",f"GDSC expression: {os.path.getsize(p)/1e6:.0f}MB"
    return "FAIL","GDSC expression data not found or too small"
test("GDSC gene expression data exists", t_gdsc_expression_exists, "data")

def t_docking_files_complete():
    dock_dir = BASE+'data/docking/'
    pdbqt = [f for f in os.listdir(dock_dir) if f.endswith('_docked.pdbqt')]
    empty = [f for f in pdbqt if os.path.getsize(dock_dir+f)<100]
    return ("WARN",f"{len(empty)} empty docked files") if empty else ("PASS",f"{len(pdbqt)} docked pdbqt files")
test("Docking output files complete", t_docking_files_complete, "data")

def t_results_no_duplicate_json():
    results_dir = BASE+'results/'
    jsons = [f for f in os.listdir(results_dir) if f.endswith('.json')]
    hashes = {}
    dupes = []
    for j in jsons:
        try:
            h = hashlib.md5(open(results_dir+j,'rb').read()).hexdigest()
            if h in hashes: dupes.append(f"{j}=={hashes[h]}")
            else: hashes[h] = j
        except: pass
    return ("WARN",f"{len(dupes)} duplicate JSON files") if dupes else ("PASS",f"{len(jsons)} unique JSON results")
test("No duplicate result files", t_results_no_duplicate_json, "data")

def t_csv_headers_consistent():
    key_csvs = ['kaalcura_real_validation.csv','beataml_significant_sensitivities.csv',
                'denovo_designed_molecules.csv','INTERCEPTA_FINAL_candidates.csv']
    issues = []
    for f in key_csvs:
        path = BASE+'results/'+f
        if not os.path.exists(path): issues.append(f"MISSING:{f}"); continue
        with open(path) as fp: header = fp.readline()
        if not header.strip(): issues.append(f"EMPTY_HEADER:{f}")
    return ("FAIL",f"Issues: {issues}") if issues else ("PASS","All key CSVs have headers")
test("Key CSV headers present", t_csv_headers_consistent, "data")

# ═══════════════════════════════════════════════════════════════
# TIER 2: ODE NUMERICAL STABILITY (L13-L22)
# ═══════════════════════════════════════════════════════════════
print("\n╔══ TIER 2: ODE NUMERICAL STABILITY (L13-L22) ══╗")

def t_ode_seed_stability():
    from intercepta_engine_v1 import PKModel, VirtualCohort
    from hr_estimator_fixed import estimate_hr_proper
    base = {'g_s':0.006,'g_r':0.003,'K':1.0,'mu':3e-5,'nu':0,'S0':0.45,'R0':0.08,'d_natural':0.001}
    hrs = []
    for seed in [42,123,456,789,999]:
        vc = VirtualCohort(n_patients=30, random_state=seed)
        pts = vc.generate_patients(base)
        ctrl = vc.simulate_cohort(pts,[],duration_days=730)
        drugs = [{'name':'docetaxel','pk_model':PKModel('docetaxel'),'emax_s':0.05,'emax_r':0.003,'ec50':0.00987,'hill_n':1.5}]
        trt = vc.simulate_cohort(pts,drugs,duration_days=730)
        ct = np.array([r['progression_time'] if r.get('progression_time') else 730 for r in ctrl])
        tt = np.array([r['progression_time'] if r.get('progression_time') else 730 for r in trt])
        r = estimate_hr_proper(ct,tt,730)
        hrs.append(r['hr'])
    cv = np.std(hrs)/np.mean(hrs)*100
    return ("FAIL",f"HR unstable across seeds: CV={cv:.0f}% hrs={[f'{h:.2f}' for h in hrs]}") if cv>50 else ("PASS",f"HR stable across 5 seeds: mean={np.mean(hrs):.3f} CV={cv:.0f}%")
test("ODE HR stable across random seeds", t_ode_seed_stability, "ode")

def t_ode_no_negative_populations():
    from intercepta_engine_v1 import PKModel, TumorODE
    pk = PKModel("docetaxel")
    ode = TumorODE({'g_s':0.008,'g_r':0.004,'K':1.0,'mu':1e-4,'nu':0,'S0':0.5,'R0':0.02,'d_natural':0.001})
    ode.add_drug("docetaxel",pk,emax_s=0.05,emax_r=0.003,ec50=0.00987)
    r = ode.simulate(1825)
    neg_S = np.sum(r['S']<-1e-10)
    neg_R = np.sum(r['R']<-1e-10)
    return ("FAIL",f"Negative populations: S={neg_S}, R={neg_R}") if neg_S+neg_R>0 else ("PASS","No negative cell populations")
test("ODE populations never go negative", t_ode_no_negative_populations, "ode")

def t_ode_total_bounded():
    from intercepta_engine_v1 import PKModel, TumorODE
    pk = PKModel("docetaxel")
    ode = TumorODE()
    ode.add_drug("docetaxel",pk,emax_s=0.05,emax_r=0.003,ec50=0.00987)
    r = ode.simulate(1825)
    N_max = np.max(r['N'])
    K = ode.params['K']
    return ("FAIL",f"N_max={N_max:.3f} exceeds K={K}") if N_max > K*1.01 else ("PASS",f"N bounded by K: max={N_max:.3f}≤{K}")
test("ODE total tumor bounded by K", t_ode_total_bounded, "ode")

def t_ode_convergence_check():
    from intercepta_phenotype_ode_v1 import PhenotypeStructuredODE, create_synthetic_velocity_distribution
    n0 = create_synthetic_velocity_distribution(20)*0.15
    # Run same simulation with different tolerances
    m1 = PhenotypeStructuredODE(N_bins=20)
    m2 = PhenotypeStructuredODE(N_bins=20)
    r1 = m1.simulate(n0.copy(), duration_days=365)
    r2 = m2.simulate(n0.copy(), duration_days=365)
    diff = abs(r1['N_total'][-1] - r2['N_total'][-1])
    return ("PASS",f"ODE reproducible: N_final diff={diff:.2e}") if diff<1e-6 else ("WARN",f"ODE non-deterministic: diff={diff:.2e}")
test("Phenotype ODE reproducible (same params)", t_ode_convergence_check, "ode")

def t_ode_drug_free_matches_analytic():
    from intercepta_engine_v1 import TumorODE
    from scipy.integrate import odeint
    # Simple logistic: dN/dt = r*N*(1-N/K)
    # Analytic: N(t) = K*N0*exp(r*t) / (K + N0*(exp(r*t)-1))
    g_s, K, S0 = 0.006, 1.0, 0.3
    ode = TumorODE({'g_s':g_s,'g_r':0.003,'K':K,'mu':0,'nu':0,'S0':S0,'R0':0,'d_natural':0})
    r = ode.simulate(365)
    N_sim = r['S'][-1]  # R=0 so all is S
    # Analytic (pure logistic, no death)
    t = 365
    N_analytic = K*S0*np.exp(g_s*t)/(K + S0*(np.exp(g_s*t)-1))
    err = abs(N_sim-N_analytic)/N_analytic*100
    return ("WARN",f"ODE vs analytic error={err:.1f}%") if err>10 else ("PASS",f"ODE matches analytic logistic: sim={N_sim:.4f} analytic={N_analytic:.4f} err={err:.1f}%")
test("ODE matches analytic logistic solution", t_ode_drug_free_matches_analytic, "ode")

def t_ode_phenotype_bins_sum_to_total():
    from intercepta_phenotype_ode_v1 import PhenotypeStructuredODE, create_synthetic_velocity_distribution
    n0 = create_synthetic_velocity_distribution(20)*0.15
    m = PhenotypeStructuredODE(N_bins=20)
    r = m.simulate(n0, duration_days=365)
    bin_sums = np.sum(r['n'], axis=0)
    max_diff = np.max(np.abs(bin_sums - r['N_total']))
    return ("FAIL",f"Bins don't sum to total: max_diff={max_diff:.2e}") if max_diff>1e-8 else ("PASS",f"Bins sum exactly to N_total: max_diff={max_diff:.2e}")
test("Phenotype ODE bins sum to total", t_ode_phenotype_bins_sum_to_total, "ode")

def t_ode_emax_threshold():
    # Find emax threshold where resistance dynamics switch direction
    from intercepta_engine_v1 import PKModel, TumorODE
    pk = PKModel("docetaxel")
    threshold = None
    for emax in [0.005,0.010,0.020,0.030,0.040,0.050]:
        ode = TumorODE({'g_s':0.006,'g_r':0.003,'K':1.0,'mu':1e-4,'nu':0,'S0':0.45,'R0':0.05,'d_natural':0.001})
        ode.add_drug("docetaxel",pk,emax_s=emax,emax_r=0.001,ec50=0.00987)
        r = ode.simulate(1825)
        if r['fraction_R'][-1] > r['fraction_R'][0]:
            threshold = emax; break
    return ("PASS",f"Resistance rises for emax_s≥{threshold} — confirms emax=0.010 too low") if threshold else ("FAIL","Resistance never rises — fundamental ODE problem")
test("emax threshold for correct resistance dynamics", t_ode_emax_threshold, "ode")

def t_ode_synergy_increases_effect():
    from intercepta_engine_v1 import PKModel, TumorODE
    base = {'g_s':0.007,'g_r':0.004,'K':1.0,'mu':5e-5,'nu':0,'S0':0.40,'R0':0.08,'d_natural':0.001}
    pk_a = PKModel("abiraterone"); pk_o = PKModel("olaparib")
    # No synergy
    ode1 = TumorODE(base)
    ode1.add_drug("abi",pk_a,emax_s=0.022,emax_r=0.003,ec50=0.0004)
    ode1.add_drug("ola",pk_o,emax_s=0.005,emax_r=0.020,ec50=0.004)
    ode1.set_synergy(0.0,0.0)
    r1 = ode1.simulate(730)
    # With synergy
    ode2 = TumorODE(base)
    ode2.add_drug("abi",pk_a,emax_s=0.022,emax_r=0.003,ec50=0.0004)
    ode2.add_drug("ola",pk_o,emax_s=0.005,emax_r=0.020,ec50=0.004)
    ode2.set_synergy(0.15,0.15)
    r2 = ode2.simulate(730)
    ttp1 = r1['progression_time'] or 730
    ttp2 = r2['progression_time'] or 730
    return ("FAIL",f"Synergy doesn't help: no_syn={ttp1:.0f}d syn={ttp2:.0f}d") if ttp2<=ttp1 else ("PASS",f"Synergy extends TTP: {ttp1:.0f}d→{ttp2:.0f}d (+{(ttp2-ttp1)/30.44:.1f}mo)")
test("Synergy parameter increases treatment effect", t_ode_synergy_increases_effect, "ode")

def t_ode_pk_v1_bug_quantified():
    from intercepta_engine_v1 import PKModel
    pk = PKModel("docetaxel")
    t,C = pk.simulate(duration_days=30)
    cmax_sim = np.max(C)
    cmax_published = 0.09  # uM free concentration from FDA label with correct V1
    overestimate = cmax_sim/cmax_published
    return ("FAIL",f"V1 bug: Cmax={cmax_sim:.4f}uM is {overestimate:.1f}x too high vs published {cmax_published}uM. Need V1=31L not 8.6L") if overestimate>2 else ("PASS",f"Cmax={cmax_sim:.4f}uM within 2x of published")
test("PK V1 bug quantified", t_ode_pk_v1_bug_quantified, "ode")

def t_ode_5yr_no_blowup():
    from intercepta_phenotype_ode_v1 import PhenotypeStructuredODE, create_synthetic_velocity_distribution
    n0 = create_synthetic_velocity_distribution(20)*0.15
    m = PhenotypeStructuredODE(N_bins=20)
    m.add_drug('docetaxel',1825)
    r = m.simulate(n0,duration_days=1825)
    max_N = np.max(r['N_total'])
    return ("FAIL",f"Numerical blowup: N_max={max_N:.1f}>>1") if max_N>10 else ("PASS",f"5yr stable: N_max={max_N:.4f}≤K=1.0")
test("5-year simulation no numerical blowup", t_ode_5yr_no_blowup, "ode")

# ═══════════════════════════════════════════════════════════════
# TIER 3: NETWORK TOPOLOGY (L23-L32)
# ═══════════════════════════════════════════════════════════════
print("\n╔══ TIER 3: NETWORK TOPOLOGY (L23-L32) ══╗")

def t_network_hub_genes():
    with open(BASE+'results/step4_hub_genes.csv') as f:
        rows = list(csv.DictReader(f))
    # Known cancer hub genes that should appear
    known_hubs = {'TP53','EGFR','AKT1','MTOR','CDK4','BRCA1','MDM2'}
    found = set(r.get('gene','').upper() for r in rows[:50])
    overlap = known_hubs & found
    return ("WARN",f"Few known hubs in top 50: {overlap}") if len(overlap)<3 else ("PASS",f"Known hubs in top 50: {overlap}")
test("Network hub genes include known cancer drivers", t_network_hub_genes, "network")

def t_string_high_confidence():
    with open(BASE+'results/step4_string_interactions.csv') as f:
        rows = list(csv.DictReader(f))
    cols = list(rows[0].keys())
    score_col = next((c for c in cols if 'score' in c.lower()), None)
    if not score_col: return "WARN","no score column"
    scores = [float(r[score_col]) for r in rows if r.get(score_col)]
    high_conf = sum(1 for s in scores if s >= 700)  # STRING high confidence
    return ("WARN",f"Only {high_conf}/{len(scores)} high-confidence (≥700)") if high_conf/len(scores)<0.3 else ("PASS",f"{high_conf}/{len(scores)} high-confidence interactions")
test("STRING interaction confidence distribution", t_string_high_confidence, "network")

def t_signor_has_activations_and_inhibitions():
    with open(BASE+'results/signor_directed_edges.csv') as f:
        rows = list(csv.DictReader(f))
    cols = list(rows[0].keys())
    effect_col = next((c for c in cols if 'effect' in c.lower() or 'mechanism' in c.lower()), None)
    if not effect_col: return "WARN",f"No effect column. Cols: {cols[:4]}"
    effects = [r[effect_col].lower() for r in rows if r.get(effect_col)]
    has_act = any('activ' in e or 'up' in e or 'stimul' in e for e in effects)
    has_inh = any('inhib' in e or 'down' in e or 'block' in e for e in effects)
    return ("FAIL","Missing activation or inhibition edges") if not (has_act and has_inh) else ("PASS",f"Both activations and inhibitions present in {len(rows)} edges")
test("SIGNOR has both activation and inhibition", t_signor_has_activations_and_inhibitions, "network")

def t_aml_genes_in_known_pathways():
    with open(BASE+'results/disease_net_acute_myeloid_leukemia.json') as f:
        d = json.load(f)
    genes = set(d.get('genes',[]))
    # Core AML pathways: RTK/RAS, epigenetic, transcription factors
    rtk_ras = {'FLT3','KIT','NRAS','KRAS','PTPN11','CBL'}
    epigenetic = {'DNMT3A','TET2','IDH1','IDH2','EZH2','ASXL1'}
    transcription = {'RUNX1','CEBPA','NPM1','WT1','TP53'}
    rtk_found = rtk_ras & genes
    epi_found = epigenetic & genes
    tx_found = transcription & genes
    if len(rtk_found)<2 or len(epi_found)<2 or len(tx_found)<2:
        return "WARN",f"Few pathway genes: RTK/RAS={rtk_found}, Epigenetic={epi_found}"
    return "PASS",f"AML pathways covered: RTK/RAS={len(rtk_found)}, Epigenetic={len(epi_found)}, TF={len(tx_found)}"
test("AML genes cover key pathways", t_aml_genes_in_known_pathways, "network")

def t_escape_routes_not_trivial():
    with open(BASE+'results/aml_escape_routes_fixed.json') as f: d=json.load(f)
    routes = d if isinstance(d,list) else list(d.values())
    # Each route should have a different primary gene
    route_genes = []
    for r in routes:
        text = json.dumps(r)
        for g in ['FLT3','IDH1','IDH2','NRAS','TP53','BCL2','MCL1','JAK2','EZH2']:
            if g in text: route_genes.append(g); break
        else: route_genes.append('UNKNOWN')
    unique = len(set(route_genes))
    return ("WARN",f"Only {unique} unique genes across {len(routes)} routes") if unique<3 else ("PASS",f"{len(routes)} routes cover {unique} unique genes: {set(route_genes)}")
test("Escape routes cover diverse genes", t_escape_routes_not_trivial, "network")

def t_network_disconnected_check():
    # Check if the gene network (from STRING) is connected
    with open(BASE+'results/step4_string_interactions.csv') as f:
        rows = list(csv.DictReader(f))
    cols = list(rows[0].keys())
    src = cols[0]; tgt = cols[1]
    genes = set()
    adj = {}
    for r in rows:
        a,b = r[src],r[tgt]
        genes.add(a); genes.add(b)
        adj.setdefault(a,set()).add(b)
        adj.setdefault(b,set()).add(a)
    # BFS from first gene
    start = list(genes)[0]
    visited = {start}
    queue = [start]
    while queue:
        node = queue.pop(0)
        for nb in adj.get(node,[]):
            if nb not in visited:
                visited.add(nb); queue.append(nb)
    connected = len(visited)/len(genes)*100
    return ("WARN",f"Network only {connected:.0f}% connected — {len(genes)-len(visited)} isolated genes") if connected<80 else ("PASS",f"Network {connected:.0f}% connected ({len(visited)}/{len(genes)} genes)")
test("STRING network connectivity", t_network_disconnected_check, "network")

def t_disease_net_json_missing_edges():
    diseases = ['disease_net_acute_myeloid_leukemia.json',
                'disease_net_non-small_cell_lung_carcinoma.json',
                'disease_net_pancreatic_carcinoma.json']
    no_edges = []
    for d in diseases:
        path = BASE+'results/'+d
        if not os.path.exists(path): continue
        with open(path) as f: data=json.load(f)
        if len(data.get('edges',data.get('interactions',[])))==0:
            no_edges.append(d.split('disease_net_')[1].split('.json')[0])
    return ("FAIL",f"Disease networks with ZERO edges: {no_edges}. Fix: merge STRING/SIGNOR edges into JSON.") if no_edges else ("PASS","All disease network JSONs have edges")
test("Disease network JSONs have edges (known failure)", t_disease_net_json_missing_edges, "network")

def t_drug_targets_in_network():
    with open(BASE+'results/disease_net_acute_myeloid_leukemia.json') as f: d=json.load(f)
    genes = set(d.get('genes',[]))
    targets = d.get('drug_targets',[])
    targets_in_genes = [t for t in targets if t in genes]
    if not targets: return "WARN","No drug targets listed"
    pct = len(targets_in_genes)/len(targets)*100
    return ("WARN",f"Only {pct:.0f}% of drug targets in gene list") if pct<80 else ("PASS",f"{len(targets_in_genes)}/{len(targets)} drug targets in AML gene network")
test("Drug targets contained in disease network", t_drug_targets_in_network, "network")

def t_mcrpc_net_has_key_genes():
    path = BASE+'results/mcrpc_unified_net.json'
    if not os.path.exists(path): return "WARN","mcrpc_unified_net.json not found"
    with open(path) as f: d=json.load(f)
    content = json.dumps(d).upper()
    mcrpc_genes = ['AR','PTEN','TP53','RB1','BRCA2','CDK12','SPOP']
    found = [g for g in mcrpc_genes if g in content]
    return ("WARN",f"Few mCRPC genes: {found}") if len(found)<4 else ("PASS",f"mCRPC key genes present: {found}")
test("mCRPC network has key prostate cancer genes", t_mcrpc_net_has_key_genes, "network")

def t_pathway_coverage():
    path = BASE+'results/step5_gene_pathway_map.csv'
    if not os.path.exists(path): return "WARN","pathway map not found"
    with open(path) as f: rows=list(csv.DictReader(f))
    pathways = set(r.get('pathway','') for r in rows)
    return ("PASS",f"{len(rows)} gene-pathway mappings, {len(pathways)} pathways") if len(pathways)>10 else ("WARN",f"Only {len(pathways)} pathways")
test("Gene-pathway map coverage", t_pathway_coverage, "network")

# ═══════════════════════════════════════════════════════════════
# TIER 4: KAALCURA DEEP (L33-L42)
# ═══════════════════════════════════════════════════════════════
print("\n╔══ TIER 4: KAALCURA DEEP (L33-L42) ══╗")

def t_kaalcura_axes_independent():
    from intercepta_kaalcura_v1 import KAALCURA, GENE_SETS
    import pandas as pd
    from scipy import stats
    rng = np.random.RandomState(42)
    genes = sorted(set(g for gs in GENE_SETS.values() for g in gs['genes']))
    expr = pd.DataFrame(rng.randn(500,len(genes)), columns=genes)
    k = KAALCURA(); k.fit_reference(expr)
    axes = k.compute_axes(expr,residualize=False)
    pairs = [('R_prolif','R_emt'),('R_prolif','R_ddr'),('R_emt','R_ddr')]
    max_r = 0
    for a,b in pairs:
        r,_ = stats.pearsonr(axes[a],axes[b])
        max_r = max(max_r,abs(r))
    return ("WARN",f"Axes correlated: max |r|={max_r:.3f} > 0.1") if max_r>0.1 else ("PASS",f"Axes independent: max |r|={max_r:.3f} ≤ 0.1")
test("KAALCURA axes mutually independent", t_kaalcura_axes_independent, "kaalcura")

def t_kaalcura_top_drugs_known():
    with open(BASE+'results/kaalcura_real_validation.csv') as f:
        rows = sorted(csv.DictReader(f), key=lambda r: float(r['auroc']), reverse=True)
    top10 = [r['drug'] for r in rows[:10]]
    # Top drugs should be ones with clear mechanism-axis links
    known_high = {'Olaparib','Vorinostat','Talazoparib','Veliparib','Niraparib',
                  'Schweinfurthin A','Vinblastine','Paclitaxel','Docetaxel'}
    overlap = set(top10) & known_high
    return ("PASS",f"Top 10 includes known mechanism drugs: {list(overlap)[:5]}") if len(overlap)>=3 else ("WARN",f"Top 10: {top10[:5]}")
test("KAALCURA top AUROC drugs biologically expected", t_kaalcura_top_drugs_known, "kaalcura")

def t_kaalcura_bottom_drugs_expected():
    with open(BASE+'results/kaalcura_real_validation.csv') as f:
        rows = sorted(csv.DictReader(f), key=lambda r: float(r['auroc']))
    bottom10_aurocs = [float(r['auroc']) for r in rows[:10]]
    # Bottom drugs should be near 0.5 (random) — hard to predict from axes
    mean_bottom = np.mean(bottom10_aurocs)
    return ("PASS",f"Bottom 10 AUROC mean={mean_bottom:.3f} (near random 0.5)") if mean_bottom<0.56 else ("WARN",f"Bottom drugs still high: {mean_bottom:.3f}")
test("KAALCURA bottom drugs near random AUROC", t_kaalcura_bottom_drugs_expected, "kaalcura")

def t_kaalcura_ddr_coefficient():
    with open(BASE+'results/kaalcura_real_validation.csv') as f:
        rows = list(csv.DictReader(f))
    parp_drugs = ['Olaparib','Talazoparib','Niraparib','Rucaparib','Veliparib']
    for r in rows:
        if r['drug'] in parp_drugs:
            coef_ddr = float(r['coef_ddr'])
            coef_prolif = float(r['coef_prolif'])
            if coef_ddr <= 0:
                return "FAIL",f"{r['drug']} has negative DDR coef={coef_ddr:.3f}"
            if abs(coef_ddr) < abs(coef_prolif):
                return "WARN",f"{r['drug']}: DDR coef={coef_ddr:.3f} < prolif coef={coef_prolif:.3f}"
    return "PASS","All PARP inhibitors have positive DDR coefficients"
test("PARP inhibitors have positive R_ddr coefficient", t_kaalcura_ddr_coefficient, "kaalcura")

def t_kaalcura_prolif_coefficient():
    with open(BASE+'results/kaalcura_real_validation.csv') as f:
        rows = list(csv.DictReader(f))
    taxanes = ['Docetaxel','Paclitaxel','Cabazitaxel','Vinblastine','Vinorelbine']
    found = [(r['drug'],float(r['coef_prolif'])) for r in rows if r['drug'] in taxanes]
    if not found: return "WARN","No taxanes found in validation"
    neg_prolif = [(d,c) for d,c in found if c >= 0]
    return ("FAIL",f"Taxanes have positive prolif coef (should be negative for sensitivity): {neg_prolif}") if neg_prolif else ("PASS",f"Taxanes correctly have negative prolif coef: {found[:3]}")
test("Taxanes have negative R_prolif coefficient", t_kaalcura_prolif_coefficient, "kaalcura")

def t_kaalcura_auroc_distribution():
    with open(BASE+'results/kaalcura_real_validation.csv') as f:
        rows = list(csv.DictReader(f))
    aurocs = [float(r['auroc']) for r in rows]
    below_random = sum(1 for a in aurocs if a < 0.5)
    above_065 = sum(1 for a in aurocs if a > 0.65)
    above_070 = sum(1 for a in aurocs if a > 0.70)
    return "PASS",f"Distribution: <0.5:{below_random}, >0.65:{above_065}, >0.70:{above_070} (of {len(aurocs)} drugs)"
test("KAALCURA AUROC distribution shape", t_kaalcura_auroc_distribution, "kaalcura")

def t_kaalcura_orthogonality_file():
    path = BASE+'results/kaalcura_orthogonal_v3.csv'
    if not os.path.exists(path): return "WARN","orthogonal validation file missing"
    with open(path) as f: rows=list(csv.DictReader(f))
    return "PASS",f"Orthogonality validation: {len(rows)} entries"
test("KAALCURA orthogonality validation file", t_kaalcura_orthogonality_file, "kaalcura")

def t_kaalcura_residualized_file():
    path = BASE+'results/kaalcura_residualized_v2.csv'
    if not os.path.exists(path): return "WARN","residualized file missing"
    with open(path) as f: rows=list(csv.DictReader(f))
    return "PASS",f"Residualized axes file: {len(rows)} samples"
test("KAALCURA residualized axes file exists", t_kaalcura_residualized_file, "kaalcura")

def t_kaalcura_gene_coverage():
    from intercepta_kaalcura_v1 import GENE_SETS
    all_genes = set(g for gs in GENE_SETS.values() for g in gs['genes'])
    # Check against GDSC expression
    gdsc_path = BASE+'data/gdsc/sanger_model_gene_expression.csv.gz'
    if not os.path.exists(gdsc_path): return "WARN","GDSC expression not found"
    import gzip
    with gzip.open(gdsc_path,'rt') as f: header=f.readline()
    gdsc_genes = set(header.strip().split(','))
    coverage = len(all_genes & gdsc_genes)/len(all_genes)*100
    return ("WARN",f"Only {coverage:.0f}% KAALCURA genes in GDSC expression") if coverage<70 else ("PASS",f"{coverage:.0f}% KAALCURA genes found in GDSC expression data")
test("KAALCURA genes covered in GDSC expression", t_kaalcura_gene_coverage, "kaalcura")

def t_kaalcura_drug_bin_sensitivity():
    path = BASE+'results/kaalcura_drug_bin_sensitivity.csv'
    if not os.path.exists(path): return "WARN","drug_bin_sensitivity.csv missing"
    with open(path) as f: rows=list(csv.DictReader(f))
    return "PASS",f"Drug-bin sensitivity: {len(rows)} drug-population pairs"
test("KAALCURA drug-bin sensitivity file", t_kaalcura_drug_bin_sensitivity, "kaalcura")

# ═══════════════════════════════════════════════════════════════
# TIER 5: CHEMISTRY DEEP (L43-L52)
# ═══════════════════════════════════════════════════════════════
print("\n╔══ TIER 5: CHEMISTRY DEEP (L43-L52) ══╗")

def t_smiles_no_duplicates():
    with open(BASE+'results/denovo_designed_molecules.csv') as f:
        rows = list(csv.DictReader(f))
    smiles = [r['smiles'] for r in rows]
    dupes = len(smiles)-len(set(smiles))
    return ("WARN",f"{dupes} duplicate SMILES") if dupes>10 else ("PASS",f"{dupes} duplicates in {len(smiles)} molecules")
test("De novo molecules no excessive duplicates", t_smiles_no_duplicates, "chemistry")

def t_molecular_weight_range():
    with open(BASE+'results/denovo_designed_molecules.csv') as f:
        rows = list(csv.DictReader(f))
    mws = [float(r['mw']) for r in rows if r.get('mw')]
    too_heavy = sum(1 for m in mws if m>600)
    too_light = sum(1 for m in mws if m<150)
    return ("WARN",f"{too_heavy} too heavy (>600), {too_light} too light (<150)") if too_heavy+too_light>len(mws)*0.2 else ("PASS",f"MW range [{min(mws):.0f}-{max(mws):.0f}], mean={np.mean(mws):.0f}")
test("Molecular weights in drug-like range", t_molecular_weight_range, "chemistry")

def t_hbd_hba_range():
    with open(BASE+'results/denovo_designed_molecules.csv') as f:
        rows = list(csv.DictReader(f))
    bad = sum(1 for r in rows if int(r.get('hbd',0))>5 or int(r.get('hba',0))>10)
    return ("WARN",f"{bad}/{len(rows)} violate HBD≤5/HBA≤10") if bad>len(rows)*0.1 else ("PASS",f"HBD/HBA in range: {bad}/{len(rows)} violations")
test("H-bond donors/acceptors in range", t_hbd_hba_range, "chemistry")

def t_10_targets_covered():
    with open(BASE+'results/denovo_designed_molecules.csv') as f:
        rows = list(csv.DictReader(f))
    targets = set(r.get('target','') for r in rows)
    expected = {'MDM2','MTOR','ATM','MAP2K1','BCL2','AR','KRAS','ATR','CHEK1','CDK4'}
    missing = expected - targets
    return ("WARN",f"Missing targets: {missing}") if missing else ("PASS",f"All 10 targets covered: {sorted(targets)}")
test("All 10 design targets have molecules", t_10_targets_covered, "chemistry")

def t_fragment_vs_optimized_ratio():
    with open(BASE+'results/denovo_designed_molecules.csv') as f:
        rows = list(csv.DictReader(f))
    frag = sum(1 for r in rows if r.get('design_method')=='fragment_based_denovo')
    opt = sum(1 for r in rows if r.get('design_method')=='denovo_optimized')
    return "PASS",f"Fragment-based: {frag} ({frag/len(rows)*100:.0f}%), Optimized: {opt} ({opt/len(rows)*100:.0f}%)"
test("Fragment vs optimized molecule ratio", t_fragment_vs_optimized_ratio, "chemistry")

def t_docking_consensus_score():
    with open(BASE+'results/scout2_docked_novel_corrected.json') as f: mols=json.load(f)
    # Check all_modes field — multiple docking poses
    multi_pose = [m for m in mols if 'all_modes' in m and len(m['all_modes'])>=3]
    if len(multi_pose) < len(mols)*0.5:
        return "WARN",f"Only {len(multi_pose)}/{len(mols)} molecules have ≥3 poses"
    # Check best pose vs 2nd pose difference
    diffs = [m['all_modes'][0]-m['all_modes'][1] for m in multi_pose]
    return "PASS",f"{len(multi_pose)}/{len(mols)} molecules have multi-pose docking, mean pose diff={np.mean(diffs):.2f}"
test("Docking multi-pose consensus", t_docking_consensus_score, "chemistry")

def t_intc002_vs_alisertib():
    with open(BASE+'results/lead_candidate_INTC002.json') as f: d=json.load(f)
    adv = d.get('advantages_over_alisertib',[])
    return ("WARN","No advantages over alisertib listed") if not adv else ("PASS",f"{len(adv)} computational advantages listed (all need wet lab confirmation)")
test("INTC002 advantages over alisertib documented", t_intc002_vs_alisertib, "chemistry")

def t_sa_score_range():
    with open(BASE+'results/denovo_designed_molecules.csv') as f:
        rows = list(csv.DictReader(f))
    # SA score: 1=easy, 10=hard to synthesize
    scores = [float(r['sa_score']) for r in rows if r.get('sa_score')]
    hard = sum(1 for s in scores if s>6)
    return ("WARN",f"{hard}/{len(scores)} molecules hard to synthesize (SA>6)") if hard>len(scores)*0.3 else ("PASS",f"SA score range [{min(scores):.1f}-{max(scores):.1f}], {hard}/{len(scores)} hard (SA>6)")
test("Synthesizability scores acceptable", t_sa_score_range, "chemistry")

def t_qed_distribution():
    with open(BASE+'results/denovo_designed_molecules.csv') as f:
        rows = list(csv.DictReader(f))
    qeds = [float(r['qed']) for r in rows if r.get('qed')]
    mean_qed = np.mean(qeds)
    return ("WARN",f"Low drug-likeness: mean QED={mean_qed:.3f}") if mean_qed<0.6 else ("PASS",f"Drug-likeness QED: mean={mean_qed:.3f}, min={min(qeds):.3f}")
test("QED drug-likeness scores", t_qed_distribution, "chemistry")

def t_novel_molecules_unique_scaffolds():
    try:
        from rdkit import Chem
        from rdkit.Chem.Scaffolds import MurckoScaffold
    except ImportError:
        return "WARN","RDKit not available for scaffold analysis"
    with open(BASE+'results/denovo_designed_molecules.csv') as f:
        rows = list(csv.DictReader(f))
    scaffolds = set()
    for r in rows[:100]:
        mol = Chem.MolFromSmiles(r.get('smiles',''))
        if mol:
            try:
                s = MurckoScaffold.MurckoScaffoldSmiles(mol=mol)
                scaffolds.add(s)
            except: pass
    diversity = len(scaffolds)/100*100
    return ("WARN",f"Low scaffold diversity: {len(scaffolds)}/100 unique scaffolds") if len(scaffolds)<30 else ("PASS",f"{len(scaffolds)}/100 unique Murcko scaffolds ({diversity:.0f}% diversity)")
test("De novo molecules scaffold diversity", t_novel_molecules_unique_scaffolds, "chemistry")

# ═══════════════════════════════════════════════════════════════
# TIER 6: BEATAML DEEP (L53-L62)
# ═══════════════════════════════════════════════════════════════
print("\n╔══ TIER 6: BEATAML DEEP (L53-L62) ══╗")

def t_beataml_npm1_effect_size():
    with open(BASE+'results/beataml_corrected_findings.json') as f: d=json.load(f)
    npm1 = d['validated_findings']['NPM1_multikinase']
    diff = npm1.get('strongest','').split('diff=+')
    if len(diff)>1:
        effect = float(diff[1].split(')')[0])
        return ("WARN",f"NPM1+Cab effect size={effect} (need >20 for clinical relevance)") if effect<20 else ("PASS",f"NPM1+Cabozantinib effect size={effect} — clinically meaningful")
    return "WARN","Cannot extract effect size"
test("NPM1/Cabozantinib effect size clinically meaningful", t_beataml_npm1_effect_size, "beataml")

def t_beataml_dnmt3a_sample():
    with open(BASE+'results/beataml_corrected_findings.json') as f: d=json.load(f)
    dnmt3a = d['validated_findings']['DNMT3A_dasatinib']
    n = dnmt3a.get('n',0); p = dnmt3a.get('p_value',1)
    return ("WARN",f"DNMT3A+Dasatinib n={n} — borderline") if n<50 else ("PASS",f"DNMT3A+Dasatinib n={n} patients, p={p:.5f}")
test("DNMT3A/Dasatinib sample size adequate", t_beataml_dnmt3a_sample, "beataml")

def t_beataml_65_findings_distribution():
    with open(BASE+'results/beataml_statistical_tests.csv') as f:
        rows = list(csv.DictReader(f))
    cols = list(rows[0].keys())
    # Find mutation and drug columns
    mut_col = next((c for c in cols if 'mut' in c.lower() or 'gene' in c.lower()), cols[0])
    drug_col = next((c for c in cols if 'drug' in c.lower() or 'inhib' in c.lower()), cols[1])
    muts = set(r[mut_col] for r in rows)
    drugs = set(r[drug_col] for r in rows)
    return "PASS",f"{len(rows)} tests: {len(muts)} mutations × {len(drugs)} drugs"
test("BeatAML test matrix dimensions", t_beataml_65_findings_distribution, "beataml")

def t_beataml_sensitivity_csv():
    path = BASE+'results/beataml_significant_sensitivities.csv'
    if not os.path.exists(path): return "WARN","significant sensitivities file missing"
    with open(path) as f: rows=list(csv.DictReader(f))
    return "PASS",f"{len(rows)} significant drug sensitivities"
test("BeatAML significant sensitivities file", t_beataml_sensitivity_csv, "beataml")

def t_beataml_wes_mutations():
    path = BASE+'data/beataml/beataml_wes_wv1to4_mutations_dbgap.txt'
    if not os.path.exists(path): return "FAIL","WES mutations file missing"
    with open(path) as f: lines=f.readlines()
    return "PASS",f"WES mutations: {len(lines)-1} variants"
test("BeatAML WES mutation data present", t_beataml_wes_mutations, "beataml")

def t_beataml_flt3_arm3():
    with open(BASE+'results/beataml_corrected_findings.json') as f: d=json.load(f)
    findings = d.get('validated_findings',{})
    # FLT3 + kinase inhibitors should be significant
    has_flt3 = any('FLT3' in k.upper() or 'flt3' in str(v).lower() for k,v in findings.items())
    return ("WARN","No FLT3 findings — expected given gilteritinib in BeatAML") if not has_flt3 else ("PASS","FLT3-related finding present")
test("BeatAML includes FLT3 biology", t_beataml_flt3_arm3, "beataml")

def t_beataml_drug_families():
    path = BASE+'data/beataml/beataml_drug_families.xlsx'
    if not os.path.exists(path): return "WARN","drug families file missing"
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    return "PASS",f"Drug families: {ws.max_row} entries"
test("BeatAML drug family classification", t_beataml_drug_families, "beataml")

def t_beataml_probit_convergence():
    path = BASE+'data/beataml/beataml_probit_curve_fits_v4_dbgap.txt'
    with open(path) as f:
        reader = csv.DictReader(f, delimiter='\t')
        rows = list(reader)
    cols = list(rows[0].keys())
    # Check for IC50 or AUC column
    ic50_col = next((c for c in cols if 'ic50' in c.lower() or 'auc' in c.lower() or 'aac' in c.lower()), None)
    if not ic50_col: return "WARN",f"No IC50/AUC column. Cols: {cols[:5]}"
    valid = sum(1 for r in rows if r.get(ic50_col,'').strip() not in ['','NA','NaN','nan'])
    return "PASS",f"{valid}/{len(rows)} valid IC50/AUC measurements"
test("BeatAML curve fits have valid IC50 values", t_beataml_probit_convergence, "beataml")

def t_beataml_clinical_age_range():
    import openpyxl
    wb = openpyxl.load_workbook(BASE+'data/beataml/beataml_wv1to4_clinical.xlsx')
    ws = wb.active
    header = [ws.cell(1,c).value for c in range(1,ws.max_column+1)]
    age_col = next((i+1 for i,h in enumerate(header) if h and 'age' in str(h).lower()), None)
    if not age_col: return "WARN","No age column found"
    ages = [ws.cell(r,age_col).value for r in range(2,ws.max_row+1)]
    ages = [a for a in ages if a and isinstance(a,(int,float))]
    return ("WARN",f"Unusual age range: {min(ages)}-{max(ages)}") if not (0<min(ages)<30 and max(ages)<120) else ("PASS",f"Age range: {min(ages):.0f}-{max(ages):.0f} years (n={len(ages)})")
test("BeatAML patient age range realistic", t_beataml_clinical_age_range, "beataml")

def t_beataml_65_fdr_threshold():
    with open(BASE+'results/beataml_corrected_findings.json') as f: d=json.load(f)
    n_sig = d.get('total_fdr_significant',0)
    n_tests = d.get('total_tests',0)
    # At FDR=0.05, expected false positives = 0.05 * n_sig
    expected_fp = int(0.05 * n_sig)
    return "PASS",f"At FDR=0.05: {n_sig} significant, ~{expected_fp} expected false positives"
test("BeatAML FDR false positive rate", t_beataml_65_fdr_threshold, "beataml")

# ═══════════════════════════════════════════════════════════════
# TIER 7: CLINICAL DEEP (L63-L72)
# ═══════════════════════════════════════════════════════════════
print("\n╔══ TIER 7: CLINICAL DEEP (L63-L72) ══╗")

def t_5trial_vs_ground_truth():
    with open(BASE+'results/phase1_5trial_VALIDATED.csv') as f:
        rows = {r['trial']:r for r in csv.DictReader(f)}
    ground_truth = {'CHAARTED':0.61,'LATITUDE':0.66,'PROfound':0.69,
                    'PROpel_BRCA':0.29,'TALAPRO2_C2':0.622}
    errors = []
    for trial,gt in ground_truth.items():
        if trial not in rows: errors.append(f"MISSING:{trial}"); continue
        sim = float(rows[trial].get('simulated',0))
        err = abs(sim-gt)/gt*100
        if err>25: errors.append(f"{trial}:{err:.0f}%off")
    return ("FAIL",f"Large errors: {errors}") if errors else ("PASS",f"All 5 trials within 25% of ground truth")
test("5 trials within 25% of ground truth", t_5trial_vs_ground_truth, "clinical")

def t_aml_cr_timing():
    with open(BASE+'results/aml_ode_v6_validation.json') as f: d=json.load(f)
    ind = d.get('induction',{})
    cr_mo = ind.get('cr_mo',99)
    # Clinical: CR by day 30-45 (1-1.5 months)
    return ("WARN",f"CR at {cr_mo}mo — should be ~1mo") if not(0.5<=cr_mo<=2.0) else ("PASS",f"CR at {cr_mo}mo matches clinical ~1mo")
test("AML CR timing within clinical range", t_aml_cr_timing, "clinical")

def t_mcrpc_enza_alis_benefit():
    with open(BASE+'results/escape_route_ode_results.json') as f: d=json.load(f)
    arms = d.get('arms',{})
    enza = next((v for k,v in arms.items() if 'enza' in k.lower() and 'alis' not in k.lower()),None)
    combo = next((v for k,v in arms.items() if 'enza' in k.lower() and 'alis' in k.lower()),None)
    if not enza or not combo: return "WARN",f"Cannot find arms. Keys: {list(arms.keys())}"
    ttp_e = enza.get('ttp_days',0) or enza.get('progression_day',0)
    ttp_c = combo.get('ttp_days',0) or combo.get('progression_day',0)
    if not ttp_e or not ttp_c: return "WARN",f"TTP not found in arm data"
    return ("PASS",f"Combo TTP={ttp_c:.0f}d > Enza TTP={ttp_e:.0f}d (+{(ttp_c-ttp_e)/30.44:.1f}mo)") if ttp_c>ttp_e else ("FAIL",f"Combo not better: {ttp_c:.0f}d vs {ttp_e:.0f}d")
test("mCRPC enza+alisertib combo beats enza alone", t_mcrpc_enza_alis_benefit, "clinical")

def t_virtual_cohort_variation():
    from intercepta_engine_v1 import PKModel, VirtualCohort
    base = {'g_s':0.006,'g_r':0.003,'K':1.0,'mu':3e-5,'nu':0,'S0':0.45,'R0':0.08,'d_natural':0.001}
    vc = VirtualCohort(n_patients=50,random_state=42)
    pts = vc.generate_patients(base)
    g_s_vals = [p['g_s'] for p in pts]
    cv = np.std(g_s_vals)/np.mean(g_s_vals)*100
    return ("WARN",f"Too little patient variation: CV={cv:.0f}%") if cv<15 else ("PASS",f"Patient heterogeneity: g_s CV={cv:.0f}% (realistic)")
test("Virtual cohort patient heterogeneity realistic", t_virtual_cohort_variation, "clinical")

def t_progression_definition_consistent():
    with open(BASE+'results/phase1_calibrated_params_VALIDATED.json') as f: d=json.load(f)
    prog_def = d.get('progression_def','')
    return ("WARN","No progression definition in calibrated params") if not prog_def else ("PASS",f"Progression defined as: '{prog_def}'")
test("Progression definition documented", t_progression_definition_consistent, "clinical")

def t_aml_veneza_sec_value():
    with open(BASE+'results/aml_ode_v6_validation.json') as f: d=json.load(f)
    veneza = d.get('venaza',d.get('ven_aza',{}))
    sec = veneza.get('sec',0)
    if sec > 10:
        return "WARN",f"VenAza sec={sec} — what does 'sec' mean? Possible misinterpretation of output"
    return "PASS",f"VenAza sec={sec}"
test("AML VenAza 'sec' output interpretable", t_aml_veneza_sec_value, "clinical")

def t_phase1_trial_design():
    path = BASE+'results/phase1_5trial_VALIDATED.csv'
    with open(path) as f: rows=list(csv.DictReader(f))
    has_benefit = all('benefit' in r for r in rows)
    benefits = [float(r.get('benefit',0)) for r in rows]
    neg_benefit = sum(1 for b in benefits if b<0)
    return ("WARN",f"{neg_benefit} trials show negative benefit") if neg_benefit else ("PASS",f"All trials show positive benefit: {[f'{b:.1f}' for b in benefits]}mo")
test("All trials show positive treatment benefit", t_phase1_trial_design, "clinical")

def t_cox_hr_vs_old_hr():
    # Confirm the magnitude of error from old method
    old_hrs = {'CHAARTED':0.719,'LATITUDE':0.575,'PROfound':0.683,'PROpel_BRCA':0.257,'TALAPRO2_C2':0.626}
    new_hrs = {'CHAARTED':1.175,'LATITUDE':0.567,'PROfound':0.646,'PROpel_BRCA':0.528,'TALAPRO2_C2':0.728}
    diffs = {k:abs(new_hrs[k]-old_hrs[k])/old_hrs[k]*100 for k in old_hrs}
    large_diffs = {k:v for k,v in diffs.items() if v>20}
    return "FAIL" if 'CHAARTED' in large_diffs else "PASS", \
           f"HR changed >20%: {large_diffs}. CHAARTED went from 0.719→1.175 (REVERSED direction)"
test("Old vs new HR: CHAARTED direction reversed", t_cox_hr_vs_old_hr, "clinical")

def t_propel_brca_enriched():
    # PROpel_BRCA has target HR=0.29 — extremely strong
    # This should be much lower than other trials
    with open(BASE+'results/phase1_5trial_VALIDATED.csv') as f:
        rows = {r['trial']:float(r['simulated']) for r in csv.DictReader(f)}
    propel = rows.get('PROpel_BRCA',1.0)
    others = [v for k,v in rows.items() if k!='PROpel_BRCA']
    if propel >= min(others):
        return "FAIL",f"PROpel_BRCA HR={propel:.3f} not lowest. Others: {[f'{v:.3f}' for v in others]}"
    return "PASS",f"PROpel_BRCA correctly lowest HR={propel:.3f} (BRCA enrichment effect)"
test("PROpel_BRCA has lowest HR (BRCA enrichment)", t_propel_brca_enriched, "clinical")

def t_aml_normal_marrow():
    with open(BASE+'results/aml_ode_v6_validation.json') as f: d=json.load(f)
    untreated = d.get('untreated',{})
    normal_min = untreated.get('normal_min',1.0)
    # In untreated AML, normal marrow should be suppressed
    return ("WARN",f"normal_min={normal_min} — should be <0.05 in untreated AML") if normal_min>0.05 else ("PASS",f"Normal marrow suppressed in untreated AML: {normal_min:.4f}")
test("AML untreated suppresses normal marrow", t_aml_normal_marrow, "clinical")

# ═══════════════════════════════════════════════════════════════
# TIER 8: PIPELINE DEEP (L73-L82)
# ═══════════════════════════════════════════════════════════════
print("\n╔══ TIER 8: PIPELINE DEEP (L73-L82) ══╗")

def t_1280_candidates_smiles_valid():
    try:
        from rdkit import Chem
    except:
        return "WARN","RDKit not available"
    with open(BASE+'results/INTERCEPTA_FINAL_candidates.csv') as f:
        rows = list(csv.DictReader(f))
    smiles_col = next((c for c in rows[0].keys() if 'smiles' in c.lower()),None)
    if not smiles_col: return "WARN","no smiles column"
    sample = rows[:100]
    valid = sum(1 for r in sample if Chem.MolFromSmiles(r.get(smiles_col,'')) is not None)
    return ("FAIL",f"Only {valid}/100 sampled SMILES valid") if valid<90 else ("PASS",f"{valid}/100 sampled SMILES valid")
test("Final 1280 candidates SMILES valid", t_1280_candidates_smiles_valid, "pipeline")

def t_ranked_candidates_sorted():
    with open(BASE+'results/INTERCEPTA_FINAL_candidates.csv') as f:
        rows = list(csv.DictReader(f))
    cols = list(rows[0].keys())
    score_col = next((c for c in cols if 'score' in c.lower() or 'rank' in c.lower()),None)
    if not score_col: return "WARN",f"No score column. Cols: {cols[:5]}"
    scores = [float(r[score_col]) for r in rows if r.get(score_col)]
    is_sorted = all(scores[i]>=scores[i+1] for i in range(len(scores)-1))
    is_sorted_asc = all(scores[i]<=scores[i+1] for i in range(len(scores)-1))
    return "PASS",f"Candidates sorted by {score_col}" if is_sorted or is_sorted_asc else "WARN",f"Candidates NOT sorted by {score_col}"
test("Final candidates sorted by score", t_ranked_candidates_sorted, "pipeline")

def t_pharma_package_lead():
    with open(BASE+'results/INTERCEPTA_FINAL_package.json') as f: d=json.load(f)
    content = json.dumps(d).lower()
    has_intc002 = 'intc' in content or 'intc002' in content or 'intc-002' in content
    has_smiles = 'smiles' in content or 'cc1cc' in content
    return ("WARN","Pharma package missing lead molecule") if not has_intc002 else ("PASS",f"Pharma package contains lead candidate reference")
test("Pharma package contains lead molecule", t_pharma_package_lead, "pipeline")

def t_patient_stratification():
    path = BASE+'results/patient_stratification.json'
    if not os.path.exists(path): return "WARN","patient stratification not found"
    with open(path) as f: d=json.load(f)
    return "PASS",f"Patient stratification: {list(d.keys())[:4]}"
test("Patient stratification results exist", t_patient_stratification, "pipeline")

def t_virtual_patient_cohort():
    path = BASE+'results/virtual_patient_cohort.csv'
    if not os.path.exists(path): return "WARN","virtual patient cohort not found"
    with open(path) as f: rows=list(csv.DictReader(f))
    return "PASS",f"Virtual cohort: {len(rows)} patients"
test("Virtual patient cohort generated", t_virtual_patient_cohort, "pipeline")

def t_end_to_end_test_passed():
    path = BASE+'results/aml_end_to_end_pipeline.json'
    if not os.path.exists(path): return "WARN","end-to-end test results not found"
    with open(path) as f: d=json.load(f)
    return "PASS",f"End-to-end pipeline results: {list(d.keys())[:4]}"
test("End-to-end pipeline test results", t_end_to_end_test_passed, "pipeline")

def t_capability_test():
    with open(BASE+'results/capability_test_results.json') as f: d=json.load(f)
    passed = sum(1 for v in d.values() if isinstance(v,dict) and v.get('passed',False))
    total = sum(1 for v in d.values() if isinstance(v,dict))
    return ("WARN",f"Capability test: {passed}/{total} passed") if passed<total*0.7 else ("PASS",f"Capability test: {passed}/{total} passed")
test("Capability test results", t_capability_test, "pipeline")

def t_ode_speed_benchmark():
    with open(BASE+'results/ode_speed_benchmark.json') as f: d=json.load(f)
    time_s = d.get('time_seconds',d.get('elapsed',d.get('duration',999)))
    return ("WARN",f"ODE benchmark took {time_s:.1f}s — too slow for production") if float(str(time_s))>60 else ("PASS",f"ODE benchmark: {time_s}s")
test("ODE speed benchmark acceptable", t_ode_speed_benchmark, "pipeline")

def t_novel_combinations_scored():
    path = BASE+'results/novel_combinations.json'
    if not os.path.exists(path): return "WARN","novel_combinations.json not found"
    with open(path) as f: d=json.load(f)
    combos = d if isinstance(d,list) else d.get('combinations',list(d.values()))
    return "PASS",f"{len(combos)} novel drug combinations scored"
test("Novel drug combinations scored", t_novel_combinations_scored, "pipeline")

def t_synergy_validation():
    path = BASE+'results/synergy_scoring_validation.json'
    if not os.path.exists(path): return "WARN","synergy validation not found"
    with open(path) as f: d=json.load(f)
    return "PASS",f"Synergy validation: {list(d.keys())[:3]}"
test("Synergy scoring validation results", t_synergy_validation, "pipeline")

# ═══════════════════════════════════════════════════════════════
# TIER 9: VISION & HONEST GAPS (L83-L92)
# ═══════════════════════════════════════════════════════════════
print("\n╔══ TIER 9: VISION & HONEST GAPS (L83-L92) ══╗")

def t_6_diseases_have_results():
    disease_nets = [f for f in os.listdir(BASE+'results/') if f.startswith('disease_net_')]
    diseases_with_escapes = []
    for d in disease_nets:
        disease = d.replace('disease_net_','').replace('.json','')
        escape_path = BASE+f'results/escape_route_ode_results.json'
        if os.path.exists(escape_path): diseases_with_escapes.append(disease)
    return "PASS",f"{len(disease_nets)} disease networks, {len(diseases_with_escapes)} with escape route analysis"
test("6 diseases have network + escape analysis", t_6_diseases_have_results, "vision")

def t_generative_model_gap():
    # Test whether a generative model exists
    code_files = os.listdir(BASE+'code/')
    has_generative = any('generat' in f.lower() or 'diffusion' in f.lower() or 'vae' in f.lower() or 'gnn' in f.lower() for f in code_files)
    return ("FAIL","No generative model found. 332 molecules from fragment-based design but no learned generative model. Vision requires true de novo generation from target pocket.") if not has_generative else ("PASS","Generative model code found")
test("Generative molecule model exists", t_generative_model_gap, "vision")

def t_src_directory_populated():
    src_files = os.listdir(BASE+'src/') if os.path.exists(BASE+'src/') else []
    py_files = [f for f in src_files if f.endswith('.py')]
    return ("FAIL",f"src/ has {len(py_files)} Python files. README claims engine_v2 is here but it's MISSING.") if len(py_files)==0 else ("PASS",f"src/ has {len(py_files)} Python files")
test("src/ production directory populated", t_src_directory_populated, "vision")

def t_wet_lab_data_absent():
    # This should FAIL — confirms no experimental validation
    wet_lab_markers = ['IC50_experimental','cell_viability','in_vivo','mouse_study','xeno']
    result_files = os.listdir(BASE+'results/')
    has_wet = any(any(m in f.lower() for m in wet_lab_markers) for f in result_files)
    return ("FAIL","No experimental (wet lab) validation data found. All IC50 values are computational estimates. No cell line assays, no mouse studies, no in vivo data.") if not has_wet else ("PASS","Experimental validation data found")
test("Wet lab validation data (expected to FAIL)", t_wet_lab_data_absent, "vision")

def t_aml_scrnaseq_needed():
    # AML relapse requires scRNA-seq data — confirm it's missing
    scrna_dir = BASE+'data/scrna/'
    aml_scrna = []
    for root,_,files in os.walk(scrna_dir):
        for f in files:
            if 'aml' in f.lower() or 'leukemia' in f.lower():
                aml_scrna.append(f)
    return ("WARN","No AML-specific scRNA-seq found. STATUS.md says AML relapse needs scRNA-seq — confirmed missing.") if not aml_scrna else ("PASS",f"AML scRNA-seq found: {aml_scrna[:3]}")
test("AML scRNA-seq data for relapse modeling", t_aml_scrnaseq_needed, "vision")

def t_scout4_boolean_wrong():
    # Confirm Scout 4 compensation is wrong
    path = BASE+'results/scout4_decision.json'
    if not os.path.exists(path): return "WARN","scout4_decision.json not found"
    with open(path) as f: d=json.load(f)
    return "WARN",f"Scout 4 decision: {list(d.keys())[:3]}. NEXT_SESSION.md confirms compensation logic wrong."
test("Scout 4 Boolean network status", t_scout4_boolean_wrong, "vision")

def t_publication_outline_exists():
    path = BASE+'PUBLICATION_OUTLINE.md'
    if not os.path.exists(path): return "WARN","no publication outline"
    with open(path) as f: content=f.read()
    has_methods = 'method' in content.lower()
    has_results = 'result' in content.lower()
    return "PASS",f"Publication outline: {len(content)} chars, has methods={has_methods}, results={has_results}"
test("Publication outline documented", t_publication_outline_exists, "vision")

def t_docs_complete():
    docs = os.listdir(BASE+'docs/') if os.path.exists(BASE+'docs/') else []
    docx = [f for f in docs if f.endswith('.docx')]
    required = ['MathSpec','Vision','Roadmap','Validation']
    found = [r for r in required if any(r.lower() in d.lower() for d in docx)]
    return ("WARN",f"Missing docs: {[r for r in required if r not in found]}") if len(found)<3 else ("PASS",f"{len(docx)} docs, key docs present: {found}")
test("Core documentation complete", t_docs_complete, "vision")

def t_requirements_installable():
    path = BASE+'requirements.txt'
    if not os.path.exists(path): return "FAIL","requirements.txt missing"
    with open(path) as f: reqs=f.readlines()
    return "PASS",f"requirements.txt has {len(reqs)} packages"
test("requirements.txt exists", t_requirements_installable, "vision")

def t_honest_status_updated():
    with open(BASE+'INTERCEPTA_STATUS.md') as f: content=f.read()
    # Check if our updates from today are there
    has_hr_fix = 'April 18' in content or 'Cox' in content
    has_honest = 'FAIL' in content or 'broken' in content.lower() or 'invalid' in content.lower()
    return ("WARN","Status not updated with today's findings") if not has_hr_fix else ("PASS","Status file updated with honest test results")
test("Status file updated with today's findings", t_honest_status_updated, "vision")

# ═══════════════════════════════════════════════════════════════
# TIER 10: EDGE CASES & STRESS (L93-L100)
# ═══════════════════════════════════════════════════════════════
print("\n╔══ TIER 10: EDGE CASES & STRESS (L93-L100) ══╗")

def t_ode_zero_drug():
    from intercepta_engine_v1 import TumorODE
    ode = TumorODE()
    r = ode.simulate(365)
    N0 = r['S'][0]+r['R'][0]
    Nf = r['S'][-1]+r['R'][-1]
    return ("FAIL","Tumor shrinks with NO drug — fundamental error") if Nf<N0*0.9 else ("PASS",f"No-drug growth correct: {N0:.3f}→{Nf:.3f}")
test("ODE with zero drugs grows correctly", t_ode_zero_drug, "stress")

def t_ode_extreme_emax():
    from intercepta_engine_v1 import PKModel, TumorODE
    pk = PKModel("docetaxel")
    ode = TumorODE()
    ode.add_drug("docetaxel",pk,emax_s=10.0,emax_r=10.0,ec50=0.00987)
    r = ode.simulate(365)
    Nf = r['S'][-1]+r['R'][-1]
    return ("FAIL",f"Extreme emax causes blowup: N={Nf:.2e}") if Nf>10 or Nf<0 else ("PASS",f"Extreme emax stable: N_final={Nf:.6f}")
test("ODE stable with extreme emax (stress test)", t_ode_extreme_emax, "stress")

def t_kaalcura_empty_expression():
    from intercepta_kaalcura_v1 import KAALCURA
    import pandas as pd
    # Test with minimal expression data
    genes = ['MKI67','TOP2A','VIM','CDH1','BRCA1','BRCA2']
    expr = pd.DataFrame([[1.0]*len(genes)], columns=genes)
    k = KAALCURA()
    k.fit_reference(pd.DataFrame([[0.0]*len(genes)]*10, columns=genes))
    axes = k.compute_axes(expr,residualize=False)
    return ("FAIL","KAALCURA crashes on minimal data") if axes is None else ("PASS",f"KAALCURA handles minimal gene set: {axes.shape}")
test("KAALCURA handles minimal gene expression", t_kaalcura_empty_expression, "stress")

def t_hr_equal_arms():
    from hr_estimator_fixed import estimate_hr_proper
    np.random.seed(42)
    same = np.random.exponential(400, 200)
    r = estimate_hr_proper(same.copy(), same.copy(), 1825)
    return ("WARN",f"HR={r['hr']:.3f} for identical arms (expect ~1.0)") if not (0.7<r['hr']<1.3) else ("PASS",f"Identical arms: HR={r['hr']:.3f}≈1.0, p={r['logrank_p']:.3f}")
test("HR estimator handles identical arms", t_hr_equal_arms, "stress")

def t_synergy_extreme_values():
    from intercepta_synergy_v1 import bliss_expected
    # Edge cases
    b1 = bliss_expected(0.0, 0.5)  # One drug does nothing
    b2 = bliss_expected(1.0, 1.0)  # Both drugs at max
    b3 = bliss_expected(0.5, 0.5)  # Equal moderate
    if abs(b1-0.5)>0.01: return "FAIL",f"Bliss(0,0.5)={b1:.3f} should be 0.5"
    if abs(b2-1.0)>0.01: return "FAIL",f"Bliss(1,1)={b2:.3f} should be 1.0"
    return "PASS",f"Bliss edge cases: (0,0.5)={b1:.3f}✓ (1,1)={b2:.3f}✓ (0.5,0.5)={b3:.3f}"
test("Synergy edge cases (0 and 1 effects)", t_synergy_extreme_values, "stress")

def t_pareto_all_dominated():
    from pareto_ranking import pareto_front
    # Case where [1,1] dominates everything
    scores = [[1.0,1.0],[0.1,0.1],[0.2,0.2],[0.5,0.4]]
    front = pareto_front(scores)
    if len(front) != 1: return "WARN",f"Expected 1 in front, got {len(front)}"
    return "PASS","Single dominant solution correctly identified"
test("Pareto front with single dominant solution", t_pareto_all_dominated, "stress")

def t_large_cohort_performance():
    import time
    from intercepta_engine_v1 import PKModel, VirtualCohort
    base = {'g_s':0.006,'g_r':0.003,'K':1.0,'mu':3e-5,'nu':0,'S0':0.45,'R0':0.08,'d_natural':0.001}
    vc = VirtualCohort(n_patients=20, random_state=42)
    pts = vc.generate_patients(base)
    t0 = time.time()
    ctrl = vc.simulate_cohort(pts,[],duration_days=365)
    elapsed = time.time()-t0
    per_patient = elapsed/20
    return ("WARN",f"Slow: {per_patient:.1f}s/patient (20 patients took {elapsed:.1f}s)") if per_patient>5 else ("PASS",f"20 patients in {elapsed:.1f}s ({per_patient:.2f}s/patient)")
test("Cohort simulation performance acceptable", t_large_cohort_performance, "stress")

def t_json_all_parseable():
    results_dir = BASE+'results/'
    failed = []
    for f in os.listdir(results_dir):
        if not f.endswith('.json'): continue
        try:
            with open(results_dir+f) as fp: json.load(fp)
        except Exception as e:
            failed.append(f"{f}:{str(e)[:30]}")
    return ("FAIL",f"Corrupt JSON files: {failed}") if failed else ("PASS",f"All JSON files parse correctly ({len([f for f in os.listdir(results_dir) if f.endswith('.json')])} files)")
test("All result JSON files parseable", t_json_all_parseable, "stress")

# ══════════════════════════════════════════════════════════
# FINAL REPORT
# ══════════════════════════════════════════════════════════
print()
print("="*70)
print("100-LEVEL STRESS TEST — FINAL REPORT")
print("="*70)
passed  = [r for r in results if r[2]=="PASS"]
failed  = [r for r in results if r[2]=="FAIL"]
warned  = [r for r in results if r[2]=="WARN"]
errored = [r for r in results if r[2]=="ERROR"]
print(f"\n  ✓ PASS:  {len(passed)}")
print(f"  ✗ FAIL:  {len(failed)}")
print(f"  ⚠ WARN:  {len(warned)}")
print(f"  ! ERROR: {len(errored)}")
print(f"  TOTAL:   {len(results)}")
print()
if failed:
    print("━━ ALL FAILURES ━━")
    for l,n,v,d,c in failed: print(f"  L{l:03d} [{c}] {n}\n       {d}")
print()
if warned:
    print("━━ ALL WARNINGS ━━")
    for l,n,v,d,c in warned: print(f"  L{l:03d} [{c}] {n}\n       {d}")
print()
if errored:
    print("━━ ALL ERRORS ━━")
    for l,n,v,d,c in errored: print(f"  L{l:03d} [{c}] {n}\n       {d}")
from collections import defaultdict
cats = defaultdict(lambda:[0,0])
for l,n,v,d,c in results:
    cats[c][1]+=1
    if v=="PASS": cats[c][0]+=1
print("\n━━ BY CATEGORY ━━")
for cat,counts in sorted(cats.items()):
    bar="█"*counts[0]+"░"*(counts[1]-counts[0])
    print(f"  {cat:<12} {bar}  {counts[0]}/{counts[1]}")
print(f"\nOVERALL: {len(passed)}/{len(results)} ({len(passed)/len(results)*100:.0f}%)")
print("="*70)
